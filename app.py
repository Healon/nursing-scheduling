"""
app.py  v3 — j3soon 核心引擎版
護理排班系統 Streamlit Web App
"""

import calendar
import datetime
import io
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from j3soon_bridge import UIConfig, run, SHIFT_DISPLAY

# ── 頁面設定 ────────────────────────────────────────────
st.set_page_config(
    page_title="護理排班系統",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.stButton>button { width: 100%; }
div[data-testid="stMetricValue"] { font-size: 1.4rem; }
</style>
""", unsafe_allow_html=True)


# ── Session State ───────────────────────────────────────
def init_state():
    defaults = {
        "nurses":     ["王小明", "李美玲", "陳志偉", "張雅惠",
                       "林佳欣", "黃俊豪", "吳雅婷", "蔡宗翰"],
        "seniority":  {},
        "year":       datetime.date.today().year,
        "month":      datetime.date.today().month,
        "min_day":    2,
        "min_eve":    2,
        "min_night":  1,
        "max_nights": 8,
        "requests":   {},
        "result":     None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ── Sidebar ─────────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/color/96/hospital.png", width=56)
    st.title("護理排班系統")
    st.caption("核心引擎：j3soon/nurse-scheduling")
    page = st.radio(
        "頁面",
        ["⚙️ 人員與班別設定", "🎓 年資設定", "📋 偏好與請假", "📅 產生排班"],
        label_visibility="collapsed",
    )
    st.divider()
    st.caption("🔗 [j3soon/nurse-scheduling](https://github.com/j3soon/nurse-scheduling)")


# ════════════════════════════════════════════════════════
# 頁面一：人員與班別設定
# ════════════════════════════════════════════════════════
if page == "⚙️ 人員與班別設定":
    st.header("⚙️ 人員與班別設定")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("排班月份")
        yc, mc = st.columns(2)
        with yc:
            st.session_state.year = st.number_input(
                "年份", min_value=2024, max_value=2030,
                value=st.session_state.year)
        with mc:
            st.session_state.month = st.selectbox(
                "月份", list(range(1, 13)),
                index=st.session_state.month - 1,
                format_func=lambda x: f"{x} 月")
        num_days = calendar.monthrange(
            st.session_state.year, st.session_state.month)[1]
        st.info(f"本月共 **{num_days}** 天")

    with col2:
        st.subheader("每班最少人數需求")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.session_state.min_day = st.number_input(
                "🌞 白班", min_value=1, max_value=10,
                value=st.session_state.min_day)
        with c2:
            st.session_state.min_eve = st.number_input(
                "🌆 小夜", min_value=1, max_value=10,
                value=st.session_state.min_eve)
        with c3:
            st.session_state.min_night = st.number_input(
                "🌙 大夜", min_value=1, max_value=10,
                value=st.session_state.min_night)

    st.divider()
    st.subheader("護理師名單")
    col_list, col_stat = st.columns([2, 1])

    with col_list:
        nurses_text = st.text_area(
            "每行一位護理師姓名",
            value="\n".join(st.session_state.nurses),
            height=220,
        )
        new_nurses = [n.strip() for n in nurses_text.split("\n") if n.strip()]
        removed = set(st.session_state.nurses) - set(new_nurses)
        for name in removed:
            st.session_state.seniority.pop(name, None)
            st.session_state.requests.pop(name, None)
        st.session_state.nurses = new_nurses

    with col_stat:
        n_total      = len(st.session_state.nurses)
        min_required = (st.session_state.min_day +
                        st.session_state.min_eve +
                        st.session_state.min_night)
        st.metric("人員總數", f"{n_total} 人")
        st.divider()
        if n_total < min_required:
            st.error(f"❌ 人數 ({n_total}) 少於每日最低需求 ({min_required} 人)")
        elif n_total < min_required + 2:
            st.warning("⚠️ 人數偏少，休假空間有限")
        else:
            st.success("✅ 人數與規則可行")

    st.divider()
    st.subheader("進階限制")
    st.session_state.max_nights = st.slider(
        "每月大夜上限（每人）", 2, 15,
        value=st.session_state.max_nights)

    st.info(
        "💡 以下規則由 **j3soon/nurse-scheduling** 核心引擎自動處理：  \n"
        "• 大夜 → 白班 禁止（休息不足）  \n"
        "• 小夜 → 白班 禁止（新增）  \n"
        "• 週末出勤均攤（內建 `WEEKEND` 關鍵字）  \n"
        "• 休假天數均攤  \n"
        "• 大夜班次均攤  \n"
        "• 台灣國定假日支援（`country: TW`）"
    )


# ════════════════════════════════════════════════════════
# 頁面二：年資設定（顯示用，不影響排班）
# ════════════════════════════════════════════════════════
elif page == "🎓 年資設定":
    st.header("🎓 年資設定")
    st.info(
        "年資資料用於排班表顯示與參考，**不作為排班約束條件**。  \n"
        "（j3soon 的 `qualifiedPeople` 語義為「限定某班只有這些人能排」，"
        "無法表達「每班至少一位資深者在場」，故此版本排除此約束）"
    )

    if not st.session_state.nurses:
        st.warning("請先至「人員與班別設定」填寫護理師名單。")
        st.stop()

    st.divider()
    st.subheader("各護理師年資輸入（年）")

    cols = st.columns(2)
    for i, name in enumerate(st.session_state.nurses):
        col = cols[i % 2]
        val = col.number_input(
            name,
            min_value=0.0, max_value=40.0, step=0.5,
            value=float(st.session_state.seniority.get(name, 0.0)),
            key=f"sen_{name}", format="%.1f",
        )
        st.session_state.seniority[name] = val

    st.divider()
    df_sen = pd.DataFrame([
        {"護理師": n, "年資(年)": st.session_state.seniority.get(n, 0.0)}
        for n in st.session_state.nurses
    ])
    df_sen = df_sen.sort_values("年資(年)", ascending=False).reset_index(drop=True)
    st.dataframe(df_sen, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════
# 頁面三：偏好與請假
# ════════════════════════════════════════════════════════
elif page == "📋 偏好與請假":
    st.header("📋 護理師偏好與請假設定")

    if not st.session_state.nurses:
        st.warning("請先至「人員與班別設定」填寫護理師名單。")
        st.stop()

    num_days = calendar.monthrange(
        st.session_state.year, st.session_state.month)[1]

    shift_options = ["（不指定）", "休", "白", "小夜", "大夜"]
    shift_map = {"（不指定）": None, "休": "休", "白": "白",
                 "小夜": "小夜", "大夜": "大夜"}

    selected_nurse = st.selectbox("選擇護理師", st.session_state.nurses)
    st.caption("設定班別偏好（不指定 = 系統自動安排）")

    existing     = st.session_state.requests.get(selected_nurse, [])
    existing_map = {d: s for d, s in existing}

    weekday_names = ["一", "二", "三", "四", "五", "六", "日"]
    first_wd      = calendar.weekday(
        st.session_state.year, st.session_state.month, 1)

    header_cols = st.columns(7)
    for i, wd in enumerate(weekday_names):
        color = "red" if i >= 5 else "gray"
        header_cols[i].markdown(
            f"<div style='text-align:center;font-weight:bold;color:{color}'>{wd}</div>",
            unsafe_allow_html=True)

    new_requests = {}
    day_idx = 0
    while day_idx < num_days:
        row_cols = st.columns(7)
        for col_pos in range(7):
            if day_idx == 0 and col_pos < first_wd:
                row_cols[col_pos].write("")
                continue
            if day_idx >= num_days:
                break
            d  = day_idx
            wd = calendar.weekday(
                st.session_state.year, st.session_state.month, d + 1)
            color = "#cc2200" if wd >= 5 else "#333"
            row_cols[col_pos].markdown(
                f"<div style='text-align:center;font-size:12px;color:{color}'>{d+1}</div>",
                unsafe_allow_html=True)
            cur_label = existing_map.get(d, "（不指定）")
            if cur_label not in shift_options:
                cur_label = "（不指定）"
            sel = row_cols[col_pos].selectbox(
                f"d{d}", shift_options,
                index=shift_options.index(cur_label),
                key=f"req_{selected_nurse}_{d}",
                label_visibility="collapsed")
            if sel != "（不指定）":
                new_requests[d] = shift_map[sel]
            day_idx += 1

    st.session_state.requests[selected_nurse] = [
        (d, s) for d, s in new_requests.items()]

    st.divider()
    st.subheader("請求摘要")
    rows = []
    for name in st.session_state.nurses:
        for d, s in st.session_state.requests.get(name, []):
            rows.append({"護理師": name,
                         "日期": f"{st.session_state.month}/{d+1}",
                         "偏好": s})
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     hide_index=True)
    else:
        st.info("目前無請求，系統將完全依規則排班。")

    if st.button("🗑️ 清除所有請求"):
        st.session_state.requests = {}
        st.rerun()


# ════════════════════════════════════════════════════════
# 頁面四：產生排班
# ════════════════════════════════════════════════════════
elif page == "📅 產生排班":
    st.header("📅 產生排班結果")

    if not st.session_state.nurses:
        st.warning("請先至「人員與班別設定」填寫護理師名單。")
        st.stop()

    num_days = calendar.monthrange(
        st.session_state.year, st.session_state.month)[1]

    with st.expander("📋 當前設定摘要", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("排班月份", f"{st.session_state.year}/{st.session_state.month}")
        c2.metric("護理師人數", f"{len(st.session_state.nurses)} 人")
        c3.metric("排班天數", f"{num_days} 天")
        c4.metric("大夜上限/人", f"{st.session_state.max_nights} 次")
        c1.metric("白班最少", f"{st.session_state.min_day} 人")
        c2.metric("小夜最少", f"{st.session_state.min_eve} 人")
        c3.metric("大夜最少", f"{st.session_state.min_night} 人")
        c4.metric("請假請求", f"{sum(len(v) for v in st.session_state.requests.values())} 筆")

    col_btn, col_note = st.columns([1, 2])
    with col_btn:
        run_btn = st.button("🚀 開始自動排班", type="primary",
                            use_container_width=True)
    with col_note:
        st.caption("通常 10–30 秒完成。使用 j3soon/nurse-scheduling CP-SAT 求解。")

    if run_btn:
        config = UIConfig(
            year=st.session_state.year,
            month=st.session_state.month,
            nurses=st.session_state.nurses,
            min_day=st.session_state.min_day,
            min_eve=st.session_state.min_eve,
            min_night=st.session_state.min_night,
            max_nights_per_month=st.session_state.max_nights,
            seniority=st.session_state.seniority,
            requests=st.session_state.requests,
            timeout=30,
        )
        with st.spinner("⏳ j3soon 求解中，請稍候..."):
            result = run(config)
        st.session_state.result = result

        if result.status in ("OPTIMAL", "FEASIBLE"):
            label = "最優解" if result.status == "OPTIMAL" else "可行解"
            st.success(f"✅ 排班完成！（{label}）")
        elif result.status == "INFEASIBLE":
            st.error(
                "❌ 找不到可行排班！可能原因：  \n"
                "• 每班需求人數 × 3 接近總人數（休假空間不足）  \n"
                "• 大夜上限過嚴  \n"
                "• 請假請求衝突過多"
            )
        else:
            st.warning(f"⚠️ 狀態：{result.status}，已回傳目前最佳解。")

    result = st.session_state.result
    if result is None or not result.schedule:
        st.info("請按「開始自動排班」產生排班表。")
        st.stop()

    st.divider()

    # ── 排班表 ────────────────────────────────────────
    st.subheader("📅 排班表")
    weekday_ch  = ["一", "二", "三", "四", "五", "六", "日"]
    col_labels  = []
    for d in range(num_days):
        wd = calendar.weekday(st.session_state.year,
                              st.session_state.month, d + 1)
        mark = "★" if wd >= 5 else ""
        col_labels.append(f"{d+1}{mark}({weekday_ch[wd]})")

    color_map = {
        "白":   "background-color:#D0EDE0",
        "小夜": "background-color:#B5D4F4",
        "大夜": "background-color:#CECBF6",
        "休":   "background-color:#F1EFE8",
    }

    df_sched = pd.DataFrame(
        {name: result.schedule[name] for name in result.schedule},
        index=col_labels,
    ).T

    st.dataframe(
        df_sched.style.applymap(lambda v: color_map.get(v, "")),
        use_container_width=True, height=360,
    )

    # ── 每日班別人數 ──────────────────────────────────
    st.subheader("📊 每日班別人數")
    daily = {"日期": col_labels, "白班": [], "小夜": [], "大夜": []}
    for d in range(num_days):
        day_shifts = [result.schedule[n][d] for n in result.schedule]
        daily["白班"].append(day_shifts.count("白"))
        daily["小夜"].append(day_shifts.count("小夜"))
        daily["大夜"].append(day_shifts.count("大夜"))
    st.dataframe(
        pd.DataFrame(daily).set_index("日期").T,
        use_container_width=True,
    )

    # ── 個人統計 ──────────────────────────────────────
    st.subheader("📊 個人統計")
    stats_df = pd.DataFrame(result.stats).T.reset_index()
    stats_df.columns = ["護理師", "上班天數", "白班", "小夜",
                        "大夜", "休假", "週末出勤", "年資(年)"]

    max_wk, min_wk = stats_df["週末出勤"].max(), stats_df["週末出勤"].min()
    max_of, min_of = stats_df["休假"].max(), stats_df["休假"].min()

    def hl(row):
        s = [""] * len(row)
        idx = list(stats_df.columns)
        if max_wk > min_wk:
            if row["週末出勤"] == max_wk:
                s[idx.index("週末出勤")] = "background-color:#ffd6cc"
            if row["週末出勤"] == min_wk:
                s[idx.index("週末出勤")] = "background-color:#d0ede0"
        if max_of > min_of:
            if row["休假"] == max_of:
                s[idx.index("休假")] = "background-color:#d0ede0"
            if row["休假"] == min_of:
                s[idx.index("休假")] = "background-color:#ffd6cc"
        return s

    st.dataframe(
        stats_df.style.apply(hl, axis=1),
        use_container_width=True, hide_index=True,
    )
    st.caption("🟥 紅色 = 偏高　🟩 綠色 = 偏低")

    # ── 均衡指標 ──────────────────────────────────────
    st.subheader("⚖️ 均衡指標")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("週末出勤差距", f"{max_wk - min_wk} 天")
    m2.metric("休假天數差距", f"{max_of - min_of} 天")
    m3.metric("大夜班差距",
              f"{stats_df['大夜'].max() - stats_df['大夜'].min()} 次")
    m4.metric("上班天數差距",
              f"{stats_df['上班天數'].max() - stats_df['上班天數'].min()} 天")

    # ── 請求滿足率 ────────────────────────────────────
    total_req = sum(len(v) for v in st.session_state.requests.values())
    if total_req > 0:
        satisfied = sum(
            1
            for name, reqs in st.session_state.requests.items()
            if name in result.schedule
            for d, s in reqs
            if 0 <= d < num_days and result.schedule[name][d] == s
        )
        st.metric("班別請求滿足率",
                  f"{int(satisfied / total_req * 100)}%",
                  help=f"{satisfied}/{total_req} 筆已滿足")

    # ── YAML 除錯檢視 ─────────────────────────────────
    with st.expander("🔍 檢視傳給 j3soon 的 YAML 設定", expanded=False):
        st.code(result.yaml_input, language="yaml")

    # ── 匯出 Excel ────────────────────────────────────
    st.divider()

    def to_excel(result, year, month):
        num_days = calendar.monthrange(year, month)[1]
        nurses   = list(result.schedule.keys())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月排班表"

        THIN   = Side(style="thin", color="FFBBBBBB")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CELL_COLORS = {
            "白":   "FFD0EDE0",
            "小夜": "FFB5D4F4",
            "大夜": "FFCECBF6",
            "休":   "FFF1EFE8",
        }

        # 標題
        ws.cell(1, 1, f"{year} 年 {month} 月 護理排班表")
        ws.cell(1, 1).font      = Font(bold=True, size=14)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=num_days + 2)

        # 日期標頭
        ws.cell(2, 1, "護理師")
        ws.cell(2, 1).font      = Font(bold=True)
        ws.cell(2, 1).alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 10
        wd_ch = ["一", "二", "三", "四", "五", "六", "日"]

        for d in range(num_days):
            col  = d + 2
            wd   = calendar.weekday(year, month, d + 1)
            cell = ws.cell(2, col, f"{d+1}\n({wd_ch[wd]})")
            cell.font      = Font(bold=True, size=9)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[2].height = 28
            ws.column_dimensions[get_column_letter(col)].width = 5.5
            if wd >= 5:
                cell.fill = PatternFill("solid", fgColor="FFFFF0C0")

        ws.cell(2, num_days + 2, "上班天數")
        ws.column_dimensions[
            get_column_letter(num_days + 2)].width = 8

        # 排班資料
        for r, name in enumerate(nurses):
            row = r + 3
            ws.cell(row, 1, name).alignment = \
                Alignment(horizontal="center")
            for d, label in enumerate(result.schedule[name]):
                col  = d + 2
                cell = ws.cell(row, col, label)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = BORDER
                fgc = CELL_COLORS.get(label, "FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor=fgc)
            ws.cell(row, num_days + 2,
                    result.stats[name]["上班天數"]
                    ).alignment = Alignment(horizontal="center")

        # 每日人數統計
        stat_row = len(nurses) + 3
        for label, offset in [("白班人數", 0),
                               ("小夜人數", 1),
                               ("大夜人數", 2)]:
            ws.cell(stat_row + offset, 1, label
                    ).font = Font(bold=True, size=9)
        labels = ["白", "小夜", "大夜"]
        for d in range(num_days):
            col = d + 2
            for i, lbl in enumerate(labels):
                cnt = sum(
                    1 for n in nurses
                    if result.schedule[n][d] == lbl)
                ws.cell(stat_row + i, col, cnt
                        ).alignment = Alignment(horizontal="center")

        # 個人統計分頁
        ws2 = wb.create_sheet("個人統計")
        headers = ["護理師", "上班天數", "白班", "小夜",
                   "大夜", "休假", "週末出勤", "年資(年)"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(1, c, h)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r, name in enumerate(nurses, 2):
            s = result.stats[name]
            for c, key in enumerate(headers, 1):
                ws2.cell(r, c,
                         name if key == "護理師" else s.get(key, "")
                         ).alignment = Alignment(horizontal="center")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    excel_bytes = to_excel(
        result, st.session_state.year, st.session_state.month)
    filename = (f"排班表_"
                f"{st.session_state.year}"
                f"{st.session_state.month:02d}.xlsx")
    st.download_button(
        "📥 下載 Excel 排班表",
        data=excel_bytes,
        file_name=filename,
        mime=("application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet"),
        type="primary",
        use_container_width=True,
    )
