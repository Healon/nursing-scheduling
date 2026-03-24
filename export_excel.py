"""
export_excel.py
將排班結果匯出為 Excel（護理長可直接使用的格式）
"""

import io
import calendar
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from scheduler import SHIFT_NAMES, SHIFT_COLORS, ScheduleResult


# 班別填色（用 openpyxl ARGB）
CELL_COLORS = {
    "休":   "FFF1EFE8",
    "白":   "FFD0EDE0",
    "小夜": "FFB5D4F4",
    "大夜": "FFCECBF6",
}

THIN = Side(style="thin", color="FFAAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def to_excel(result: ScheduleResult, year: int, month: int) -> bytes:
    """
    產生 Excel 的 bytes，可直接用 st.download_button 提供下載。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月排班表"

    num_days = calendar.monthrange(year, month)[1]
    nurses   = list(result.schedule.keys())

    # ── 標題列 ────────────────────────────────────────
    ws.cell(1, 1, f"{year} 年 {month} 月 護理排班表")
    ws.cell(1, 1).font      = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 2)

    # ── 日期標頭 ──────────────────────────────────────
    ws.cell(2, 1, "護理師")
    ws.cell(2, 1).font      = Font(bold=True)
    ws.cell(2, 1).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 10

    weekday_names = ["一", "二", "三", "四", "五", "六", "日"]
    for d in range(num_days):
        col   = d + 2
        dt    = calendar.weekday(year, month, d + 1)
        label = f"{d+1}\n({weekday_names[dt]})"
        cell  = ws.cell(2, col, label)
        cell.font      = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30
        ws.column_dimensions[get_column_letter(col)].width = 5.5

        # 週末底色
        if dt >= 5:
            cell.fill = PatternFill("solid", fgColor="FFFFF0C0")

    ws.cell(2, num_days + 2, "上班天數")
    ws.column_dimensions[get_column_letter(num_days + 2)].width = 8

    # ── 排班資料 ──────────────────────────────────────
    for r, name in enumerate(nurses):
        row = r + 3
        ws.cell(row, 1, name).alignment = Alignment(horizontal="center")

        shifts = result.schedule[name]
        for d, sid in enumerate(shifts):
            label = SHIFT_NAMES[sid]
            col   = d + 2
            cell  = ws.cell(row, col, label)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER
            color = CELL_COLORS.get(label, "FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)

        # 上班天數統計
        work_days = result.stats[name]["上班天數"]
        ws.cell(row, num_days + 2, work_days).alignment = Alignment(horizontal="center")

    # ── 每日人數統計列 ────────────────────────────────
    stat_row = len(nurses) + 3
    ws.cell(stat_row, 1, "白班人數").font = Font(bold=True, size=9)
    ws.cell(stat_row + 1, 1, "小夜人數").font = Font(bold=True, size=9)
    ws.cell(stat_row + 2, 1, "大夜人數").font = Font(bold=True, size=9)

    from scheduler import SHIFT_DAY, SHIFT_EVE, SHIFT_NIGHT
    for d in range(num_days):
        col = d + 2
        day_counts = {
            SHIFT_DAY:   0,
            SHIFT_EVE:   0,
            SHIFT_NIGHT: 0,
        }
        for name in nurses:
            sid = result.schedule[name][d]
            if sid in day_counts:
                day_counts[sid] += 1

        ws.cell(stat_row,     col, day_counts[SHIFT_DAY]).alignment   = Alignment(horizontal="center")
        ws.cell(stat_row + 1, col, day_counts[SHIFT_EVE]).alignment   = Alignment(horizontal="center")
        ws.cell(stat_row + 2, col, day_counts[SHIFT_NIGHT]).alignment = Alignment(horizontal="center")

    # ── 統計分頁 ──────────────────────────────────────
    ws2 = wb.create_sheet("個人統計")
    headers = ["護理師", "上班天數", "白班", "小夜", "大夜", "休假"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(1, c, h)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r, name in enumerate(nurses, 2):
        s = result.stats[name]
        ws2.cell(r, 1, name)
        ws2.cell(r, 2, s["上班天數"])
        ws2.cell(r, 3, s["白班"])
        ws2.cell(r, 4, s["小夜"])
        ws2.cell(r, 5, s["大夜"])
        ws2.cell(r, 6, s["休假"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
